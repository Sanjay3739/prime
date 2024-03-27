import streamlit as st
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
from langchain_openai import OpenAIEmbeddings
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook


def get_excel_data(excel):
    data = []
    try:
        workbook = load_workbook(filename=excel)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_data = {sheet_name: [str(cell) if cell is not None else "" for cell in row]}
                data.append(row_data)
        st.success("Successfully read Excel file.")
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
    return data 


# Function to process PDF files
def get_pdf_text(pdf):
    text = ""
    try:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            if page.extract_text() is not None:
                text += page.extract_text()
                st.success(f"Successfully uploaded files")
    except Exception as e:
        st.error(f"Error reading PDF file: {e}")
    return text


# Function to handle the uploaded files based on their types
def process_uploaded_files(uploaded_files):
    dataset = []
    for uploaded_file in uploaded_files:
        if uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            dataset += get_excel_data(uploaded_file)
            print("data seat", dataset)
    return dataset


# Function to split text into chunks
def get_text_chunks(raw_data):
    text_chunks = []
    for item in raw_data:
        for sheet_name, rows in item.items():
            for row in rows:
                text_chunks.append("\t".join([str(cell) for cell in row]))
    return text_chunks


# Function to create a vector store
def get_vectorStore(text_chunks):
    if not text_chunks:  # Check if text_chunks is empty
        st.error(
            "No text found for processing. Please ensure the uploaded files contain text."
        )
        return None
    try:
        embeddings = OpenAIEmbeddings()
        vectorStore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
        return vectorStore
    except Exception as e:
        st.error(f"Error creating vector store: {e}")
        return None


# Function to create a conversation chain
def get_conversation_chain(vectorStore):
    llm = ChatOpenAI()
    memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm, retriever=vectorStore.as_retriever(), memory=memory
    )
    return conversation_chain


# Function to handle user input
def handle_userinput(user_question):
    try:
        if st.session_state.conversation is not None:
            response = st.session_state.conversation({"question": user_question})
            st.session_state.chat_history = response["chat_history"]
            for i, message in enumerate(st.session_state.chat_history):
                if i % 2 == 0:
                    st.write(
                        user_template.replace("{{MSG}}", message.content),
                        unsafe_allow_html=True,
                    )
                else:
                    st.write(
                        bot_template.replace("{{MSG}}", message.content),
                        unsafe_allow_html=True,
                    )
        else:
            st.error(
                "The conversation model is not initialized. Please process the PDF files first."
            )
    except (
        Exception
    ) as e:  # Replace 'Exception' with the specific exception type from the API library
        if (
            "rate_limit" in str(e).lower()
        ):  # Check if the exception message is about rate limiting
            st.error("We have reached the API's rate limit. Please try again later.")
        else:
            st.error(f"An error occurred: {e}")


# Main function
def main():
    load_dotenv()
    st.set_page_config(
        page_title="Chat with multiple PDFs & Excel", page_icon=":books:"
    )
    st.write(css, unsafe_allow_html=True)

    # Initialize session state variables
    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None
    if "pdf_processed" not in st.session_state:
        st.session_state.pdf_processed = False

    st.header("Chat with multiple PDFs & Excel :books:")

    # Sidebar for uploading PDF documents
    with st.sidebar:
        st.subheader("Your documents")
        uploaded_files = st.file_uploader(
            "Upload your Excel and PDF files here and click on 'Process'",
            accept_multiple_files=True,
            type=["xlsx", "pdf"],
        )
        if st.button("Process"):
            if not uploaded_files:
                st.error("Please upload your files to get started.")
            else:
                with st.spinner("Processing"):

                    # Process uploaded files
                    raw_text = process_uploaded_files(uploaded_files)
                    # Get the text chunks
                    text_chunks = get_text_chunks(raw_text)

                    # Create vector store only if there are text chunks
                    if text_chunks:
                        vectorStore = get_vectorStore(text_chunks)

                        if vectorStore is not None:
                            # Create conversation chain
                            st.session_state.conversation = get_conversation_chain(
                                vectorStore
                            )
                            st.session_state.pdf_processed = True
                        else:
                            st.error(
                                "Failed to create vector store from the provided text."
                            )
                    else:
                        st.error("No text extracted from the files for processing.")

    # Text input for user questions
    user_question = st.text_input("Ask a question about your documents:")
    if user_question:
        if st.session_state.pdf_processed:
            handle_userinput(user_question)
        else:
            st.error("Before proceeding, please upload your files.")


# Run the main function
if __name__ == "__main__":
    main()
