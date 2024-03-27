# Import statements
import streamlit as st
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from htmlTemplates import css, bot_template, user_template
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
from langchain_openai import OpenAIEmbeddings
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
# Load environment variables
load_dotenv()


# Function to process Excel files
def get_excel_text(excel):
    text = ""
    try:
        workbook = load_workbook(filename=excel)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
                values_only=True,
            ):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
    return text


# Function to process PDF files
def get_pdf_text(pdf):
    text = ""
    try:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            if page.extract_text() is not None:
                text += page.extract_text()
    except Exception as e:
        st.error(f"Error reading PDF file: {e}")
    return text


# Function to handle uploaded files
def process_uploaded_files(uploaded_files):
    raw_text = ""
    for uploaded_file in uploaded_files:
        if uploaded_file.type == "application/pdf":
            raw_text += get_pdf_text(uploaded_file)
        elif (
            uploaded_file.type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            raw_text += get_excel_text(uploaded_file)
    return raw_text


def handle_userinput(user_question):
    try:
        if st.session_state.conversation is not None:
            response = st.session_state.conversation({"question": user_question})
            st.session_state.chat_history = response["chat_history"]
            for i, message in enumerate(st.session_state.chat_history):
                if i % 2 == 0:
                    st.write(
                        user_template.replace("{{MSG}}", message.content),
                        unsafe_allow_html=True,
                    )
                else:
                    st.write(
                        bot_template.replace("{{MSG}}", message.content),
                        unsafe_allow_html=True,
                    )
        else:
            st.error(
                "The conversation model is not initialized. Please process the PDF files first."
            )
    except (
        Exception
    ) as e:  # Replace 'Exception' with the specific exception type from the API library
        if (
            "rate_limit" in str(e).lower()
        ):  # Check if the exception message is about rate limiting
            st.error("We have reached the API's rate limit. Please try again later.")
        else:
            st.error(f"An error occurred: {e}")
def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n", chunk_size=1000, chunk_overlap=200, length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks

def get_vectorStore(text_chunks):
    if not text_chunks:  # Check if text_chunks is empty
        st.error(
            "No text found for processing. Please ensure the uploaded files contain text."
        )
        return None
    try:
        embeddings = OpenAIEmbeddings()
        vectorStore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
        return vectorStore
    except Exception as e:
        st.error(f"Error creating vector store: {e}")
        return None
    
def get_conversation_chain(vectorStore):
    llm = ChatOpenAI()
    memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm, retriever=vectorStore.as_retriever(), memory=memory
    )
    return conversation_chain    
# Main function
def main():
    load_dotenv()
    st.set_page_config(
        page_title="Chat with multiple PDFs & Excel", page_icon=":books:"
    )
    st.write(css, unsafe_allow_html=True)

    # Initialize session state variables
    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None
    if "pdf_processed" not in st.session_state:
        st.session_state.pdf_processed = False

    st.header("Chat with multiple PDFs & Excel :books:")
with st.sidebar:
    uploaded_files = st.file_uploader("Upload your files", accept_multiple_files=True)

    if st.button("Process"):
        if not uploaded_files:
            st.error("Please upload files to get started.")
        else:
            with st.spinner("Processing files..."):
                raw_text = process_uploaded_files(uploaded_files)
                if raw_text:
                    st.success("Files processed successfully.")
                    # You can add chat logic here to respond to the user's question
                else:
                    st.error("No text found in the uploaded files.")
                text_chunks = get_text_chunks(raw_text)

                # Create vector store only if there are text chunks
                if text_chunks:
                    vectorStore = get_vectorStore(text_chunks)

                    if vectorStore is not None:
                        # Create conversation chain
                        st.session_state.conversation = get_conversation_chain(
                            vectorStore
                        )
                        st.session_state.pdf_processed = True
                    else:
                        st.error(
                            "Failed to create vector store from the provided text."
                        )
                else:
                    st.error("No text extracted from the files for processing.")

    # Text input for user questions
    user_question = st.text_input("Ask a question about your documents:")
    if user_question:
        if st.session_state.pdf_processed:
            handle_userinput(user_question)
        else:
            st.error("Before proceeding, please upload your files.")


if __name__ == "__main__":
    main()
